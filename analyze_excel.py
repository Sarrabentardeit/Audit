import pandas as pd
import openpyxl
import json

file_path = 'Grille 2025 (1).xlsx'

# Charger le classeur
wb = openpyxl.load_workbook(file_path, data_only=True)

print("=" * 80)
print("ANALYSE DU FICHIER EXCEL")
print("=" * 80)

# Analyser chaque feuille
for sheet_name in wb.sheetnames:
    print(f"\n{'='*80}")
    print(f"FEUILLE: {sheet_name}")
    print(f"{'='*80}")
    
    ws = wb[sheet_name]
    
    # Afficher les dimensions
    print(f"Dimensions: {ws.max_row} lignes x {ws.max_column} colonnes")
    
    # Lire avec pandas pour voir la structure
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        # Afficher les premières lignes non vides
        print(f"\nPremières lignes (max 20):")
        non_empty_rows = df.dropna(how='all')
        print(non_empty_rows.head(20).to_string())
        
        # Chercher des patterns importants
        print(f"\nRecherche de colonnes clés...")
        for col_idx in range(min(10, df.shape[1])):
            col_data = df.iloc[:, col_idx].dropna().astype(str)
            unique_vals = col_data.unique()[:10]
            if len(unique_vals) > 0:
                print(f"  Colonne {col_idx}: {unique_vals[:5]}")
        
    except Exception as e:
        print(f"Erreur lors de la lecture pandas: {e}")
        # Afficher directement depuis openpyxl
        print("\nPremières cellules (lignes 1-20, colonnes A-J):")
        for row in range(1, min(21, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(11, ws.max_column + 1)):
                cell_value = ws.cell(row, col).value
                if cell_value is not None:
                    row_data.append(str(cell_value)[:30])
                else:
                    row_data.append("")
            if any(row_data):
                print(f"Ligne {row}: {' | '.join(row_data)}")

print("\n" + "=" * 80)
print("ANALYSE TERMINÉE")
print("=" * 80)



